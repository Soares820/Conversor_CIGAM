"""
template.py
-----------
Carrega o gabarito de uma aba do modelo CIGAM.

Cada aba do arquivo-modelo segue o padrao:
  Linha 1 (R1): nome real da coluna no banco (cabecalho)
  Linha 2 (R2): regra / tamanho do campo (ex.: "20 caracteres")
  Linha 3 (R3): linha DEFAULT — valores-padrao (espacos de largura fixa,
                'A05', 'NULL', 0, datas). E o que sera injetado nos campos
                que o cliente NAO mapear.
  Linha 4+   : exemplos ou dados ja existentes (ignorados na conversao).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl


# largura declarada em regras como "20 caracteres" / "1 caracter"
_RE_TAMANHO = re.compile(r"(\d+)\s*caracter", re.IGNORECASE)
# nome da tabela: prefere o que estiver entre parenteses; senao, ultima
# palavra em MAIUSCULAS do nome da aba (ex.: "Contas a Pagar GFLANCAM")
_RE_TBL_PAREN = re.compile(r"\(([A-Za-z0-9_]+)\)")
_RE_TBL_UPPER = re.compile(r"\b([A-Z][A-Z0-9_]{3,})\b")


def extrair_tabela(nome_aba: str) -> str:
    m = _RE_TBL_PAREN.search(nome_aba)
    if m:
        return m.group(1).upper()
    candidatos = _RE_TBL_UPPER.findall(nome_aba)
    if candidatos:
        return candidatos[-1].upper()
    return nome_aba.strip().replace(" ", "_").upper()


def _tamanho_do_campo(regra: Any) -> int | None:
    if regra is None:
        return None
    m = _RE_TAMANHO.search(str(regra))
    return int(m.group(1)) if m else None


@dataclass
class CigamTemplate:
    """Gabarito de uma aba/tabela CIGAM."""

    aba: str
    tabela: str
    colunas: list[str]
    regras: list[Any]
    defaults: list[Any]
    tamanhos: list[int | None] = field(default_factory=list)

    @classmethod
    def de_arquivo(cls, caminho_modelo: str, aba: str) -> "CigamTemplate":
        wb = openpyxl.load_workbook(caminho_modelo, data_only=True)
        if aba not in wb.sheetnames:
            raise ValueError(
                f"Aba '{aba}' nao existe no modelo. "
                f"Abas: {wb.sheetnames}"
            )
        ws = wb[aba]
        ncol = ws.max_column

        colunas = [ws.cell(1, c).value for c in range(1, ncol + 1)]
        regras = [ws.cell(2, c).value for c in range(1, ncol + 1)]
        defaults = [ws.cell(3, c).value for c in range(1, ncol + 1)]
        tamanhos = [_tamanho_do_campo(r) for r in regras]

        # remove colunas totalmente vazias no fim (ruido do xlsx)
        while colunas and colunas[-1] in (None, ""):
            colunas.pop(); regras.pop(); defaults.pop(); tamanhos.pop()

        return cls(
            aba=aba,
            tabela=extrair_tabela(aba),
            colunas=[str(c) for c in colunas],
            regras=regras,
            defaults=defaults,
            tamanhos=tamanhos,
        )

    @property
    def ncolunas(self) -> int:
        return len(self.colunas)

    def indice(self, coluna: str) -> int:
        try:
            return self.colunas.index(coluna)
        except ValueError as exc:
            raise KeyError(
                f"Coluna CIGAM '{coluna}' nao existe em {self.tabela}."
            ) from exc

    def resumo_colunas(self) -> list[dict]:
        """Util para montar a tela de De-Para: nome, regra, tamanho, default."""
        return [
            {
                "coluna": self.colunas[i],
                "regra": self.regras[i],
                "tamanho": self.tamanhos[i],
                "default": self.defaults[i],
            }
            for i in range(self.ncolunas)
        ]


def listar_abas(caminho_modelo: str) -> list[dict]:
    """Lista todas as abas do modelo com a tabela destino detectada."""
    wb = openpyxl.load_workbook(caminho_modelo, data_only=True)
    return [
        {"aba": nome, "tabela": extrair_tabela(nome)}
        for nome in wb.sheetnames
    ]
