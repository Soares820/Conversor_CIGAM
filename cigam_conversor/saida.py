"""
saida.py
--------
Gera os artefatos finais a partir de um ResultadoConversao:
  - XLSX no layout CIGAM (cabecalho R1 + dados)
  - script SQL de carga em staging (INSERT)
  - script INSERT ... SELECT de promocao staging -> tabela CIGAM
"""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from .conversor import ResultadoConversao


# ---------------------------------------------------------------- XLSX ---- #
def gerar_xlsx(resultado: ResultadoConversao, caminho: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = resultado.tabela[:31]

    ws.append(resultado.colunas)
    cabecalho_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = cabecalho_fill

    for linha in resultado.linhas:
        ws.append(linha)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    wb.save(caminho)
    return caminho


# ----------------------------------------------------------------- SQL ---- #
def _sql_valor(v: Any, nulls_reais: bool) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, (datetime, date)):
        return "'" + v.strftime("%Y-%m-%d %H:%M:%S") + "'"
    s = str(v)
    if nulls_reais and s.strip().upper() == "NULL":
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def _escrever(caminho, texto: str):
    """caminho: path (str/Path) ou objeto com .write() (ex.: io.StringIO)."""
    if hasattr(caminho, "write"):
        caminho.write(texto)
    else:
        with open(caminho, "w", encoding="utf-8") as f:
            f.write(texto)
    return caminho


def gerar_sql_staging(
    resultado: ResultadoConversao,
    caminho,
    *,
    schema: str = "dbo",
    prefixo_staging: str = "stg_",
    nulls_reais: bool = True,
    lote: int = 1000,
) -> str:
    """
    Gera INSERTs para a tabela de staging.
    nulls_reais : converte a string 'NULL' em NULL de verdade.
    lote        : agrupa em blocos (GO a cada N inserts) p/ SQL Server.
    """
    tabela = f"[{schema}].[{prefixo_staging}{resultado.tabela}]"
    cols = ", ".join(f"[{c}]" for c in resultado.colunas)

    partes: list[str] = [
        f"-- Carga staging para {resultado.tabela}",
        f"-- {len(resultado.linhas)} registro(s)",
        "SET NOCOUNT ON;",
        "",
    ]
    for n, linha in enumerate(resultado.linhas, start=1):
        vals = ", ".join(_sql_valor(v, nulls_reais) for v in linha)
        partes.append(f"INSERT INTO {tabela} ({cols}) VALUES ({vals});")
        if lote and n % lote == 0:
            partes.append("GO")

    return _escrever(caminho, "\n".join(partes) + "\n")


def gerar_sql_promocao(
    resultado: ResultadoConversao,
    caminho,
    *,
    schema: str = "dbo",
    prefixo_staging: str = "stg_",
    pk: list[str] | None = None,
) -> str:
    """
    Gera o INSERT ... SELECT que promove os dados da staging para a tabela
    CIGAM real, evitando duplicar registros ja existentes (por PK).
    """
    stg = f"[{schema}].[{prefixo_staging}{resultado.tabela}]"
    dst = f"[{schema}].[{resultado.tabela}]"
    cols = ", ".join(f"[{c}]" for c in resultado.colunas)

    linhas = [
        f"-- Promocao staging -> CIGAM ({resultado.tabela})",
        f"INSERT INTO {dst} ({cols})",
        f"SELECT {cols}",
        f"FROM {stg} s",
    ]
    if pk:
        cond = " AND ".join(f"d.[{c}] = s.[{c}]" for c in pk)
        linhas += [
            "WHERE NOT EXISTS (",
            f"    SELECT 1 FROM {dst} d WHERE {cond}",
            ");",
        ]
    else:
        linhas[-1] += ";"

    return _escrever(caminho, "\n".join(linhas) + "\n")
